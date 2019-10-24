# -*- coding: utf-8 -*-
import openpyxl

workbook = openpyxl.load_workbook('data.xlsx')

if not workbook:
    print "can not load, data.xlsx"
    exit(0)

sheet = workbook['text'] 

language = {}
for col in range(2, sheet.max_column+1):
    lang = sheet.cell(row=1, column=col).value
    print lang
    values = {}
    for row in range(2, sheet.max_row+1):
        key = sheet.cell(row=row, column=1).value
        value = sheet.cell(row=row, column=col).value
        values[key] = value
    language[lang] = values
workbook.close()

data = open('data.ts', 'w')
for lang in language:
    data.write('export const '+lang+' = {\n')
    langs = language[lang]
    for key in langs:
        if langs[key]:
            data.write('\'{0}\':`{1}`, \n'.format(key.encode('utf-8'), langs[key].encode('utf-8')))
    data.write('}\n')
data.close()